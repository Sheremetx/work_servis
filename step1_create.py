from openpyxl import Workbook
import random

coutn = 0
wb = Workbook()
ws = wb.active
ws.title = "Данные"

ws_a = []

for row in range(1,51):
    ws[f"A{row}"] = random.randint(10,998)
    ws_a.append(ws[f"A{row}"].value)
    
max_value = max(ws_a)
min_value = min(ws_a)
total = sum(ws_a)


print(max_value)
print(min_value)
print(total)



stat = wb.create_sheet("Статистика")
    
stat["A1"] = f"Максимум: {max_value}"
stat["A2"] = f"Минимум: {min_value}"
stat["A3"] = f"Сумма: {total}"
    
wb.save("домашка_1.xlsx")
